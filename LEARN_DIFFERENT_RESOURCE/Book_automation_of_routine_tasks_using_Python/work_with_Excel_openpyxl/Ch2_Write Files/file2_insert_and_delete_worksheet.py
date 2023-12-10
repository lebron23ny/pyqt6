import openpyxl
wb = openpyxl.Workbook()
print(wb.sheetnames)
wb.create_sheet()
print(wb.sheetnames)


wb.create_sheet(index=0, title='Первый лист')
print(wb.sheetnames)
wb.create_sheet(index=2, title='Средний лист')
print(wb.sheetnames)

del wb['Средний лист']
print(wb.sheetnames)
del wb['Sheet1']
print(wb.sheetnames)