

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
print(sheet.title)
max_row = sheet.max_row
print(max_row)
PRICE_UPDATES = {'Lime': 3.07, 'Celery': 1.19, 'Garlic': 1.27}

for row in range(2, max_row):
    product = sheet.cell(row=row, column=1).value
    if product in PRICE_UPDATES:
        sheet.cell(row=row, column=2).value = PRICE_UPDATES[product]


wb.save('updateProduceSales.xlsx')
