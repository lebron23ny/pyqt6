import openpyxl
wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3')

sheet.unmerge_cells('C5:E5')

wb.save('unmerged.xlsx')