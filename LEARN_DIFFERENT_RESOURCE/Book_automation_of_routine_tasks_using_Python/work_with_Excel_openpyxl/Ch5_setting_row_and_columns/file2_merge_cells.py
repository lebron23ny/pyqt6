import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Объединены двенадцать ячеек'
sheet.merge_cells('C5:E5')
sheet['C5'] = 'Объединены три ячейки'
wb.save('merged.xlsx')