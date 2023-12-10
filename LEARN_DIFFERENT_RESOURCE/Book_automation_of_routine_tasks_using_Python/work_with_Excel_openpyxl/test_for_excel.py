import openpyxl



wb = openpyxl.load_workbook('../../../My_Application/App_2/Data.xlsx')
print(type(wb))

# Получаем имена листов
sheet_name_list = wb.sheetnames
print(sheet_name_list)

sheet_I_beam = wb[('Профили_ГСП'
                   '')]
# sheet_I_beam = wb['Лист1']

print(sheet_I_beam.title)

print('Активный лист:', wb.active.title)

max_row = sheet_I_beam.max_row
max_column = sheet_I_beam.max_column

print(max_row)
print(max_column)
print()
profile_list = sheet_I_beam[f'A3:A{max_row}']
# for currentRow in profile_list:
#     for currentCell in currentRow:
#         print(currentCell.value)


for cell in sheet_I_beam.columns[1]:
    print(cell.value)

wb.close()