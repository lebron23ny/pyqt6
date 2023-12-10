from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QLabel, QComboBox, QVBoxLayout, QGridLayout,
                             QTextEdit, QPushButton)

import sys
import openpyxl

class Mainwindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.InitializeComponent()
        self.SetValuesDefault()

        self.BindingSignal()
        self.SetStyle()






    def InitializeComponent(self):
        widget = QWidget()
        main_layot = QVBoxLayout()
        grid = QGridLayout()
        self.wb = openpyxl.load_workbook('Data.xlsx')

        self.label_title = QLabel('Заглавный лейбл')
        self.label_type_profile = QLabel('Тип профиля')
        self.label_profile = QLabel('Профиль')
        self.label_steel = QLabel('Сталь')

        self.combobox_type_profile = QComboBox()
        self.combobox_profile = QComboBox()
        self.combobox_grade_steel = QComboBox()
        self.button = QPushButton('Кнопка для тестов')

        self.output = QTextEdit()
        grid.addWidget(self.label_type_profile, 0, 0)
        grid.addWidget(self.label_profile, 0, 1)
        grid.addWidget(self.label_steel, 0, 2)
        grid.addWidget(self.combobox_type_profile, 1, 0)
        grid.addWidget(self.combobox_profile, 1, 1)
        grid.addWidget(self.combobox_grade_steel, 1, 2)

        main_layot.addWidget(self.label_title)
        main_layot.addLayout(grid)
        main_layot.addWidget(self.output)
        main_layot.addWidget(self.button)

        widget.setLayout(main_layot)

        self.setCentralWidget(widget)

    def SetValuesDefault(self):
        self.combobox_type_profile.addItems(['Труба', 'Двутавр'])
        self.set_values_cb_profile(self.combobox_type_profile.currentText())
        self.set_values_cb_grade_steel(self.combobox_type_profile.currentText())

        self.get_output()



    def SetStyle(self):

        self.setMinimumSize(345, 210)
        self.setWindowTitle('Приложение №2 ver 1')


    def BindingSignal(self):
        self.button.clicked.connect(self.the_button_click)
        self.combobox_type_profile.currentTextChanged.connect(self.the_cb_type_change)
        self.combobox_profile.currentTextChanged.connect(self.the_cb_profile_change)
        self.combobox_grade_steel.currentTextChanged.connect(self.the_cb_grade_steel)




    def the_button_click(self):
        print('click')
        print(self.combobox_type_profile.currentIndex())
        print(self.combobox_type_profile.currentText())
        print()


    def the_cb_type_change(self):
        self.set_values_cb_profile(self.combobox_type_profile.currentText())
        self.set_values_cb_grade_steel(self.combobox_type_profile.currentText())


    def the_cb_profile_change(self):
        self.get_output()


    def the_cb_grade_steel(self):
        self.get_output()

    def set_values_cb_profile(self, type_profile):
        self.combobox_profile.clear()
        if type_profile == "Двутавр":


            sheet_I_beam = self.wb['Профили_Двутавры']
            max_row = sheet_I_beam.max_row
            profile_list = sheet_I_beam[f'A3:A{max_row}']
            for currentRow in profile_list:
                for currentCell in currentRow:
                    self.combobox_profile.addItem(currentCell.value)


        elif type_profile == "Труба":

            sheet_tube = self.wb['Профили_ГСП']
            max_row = sheet_tube.max_row
            profile_list = sheet_tube[f'A3:A{max_row}']
            for currentRow in profile_list:
                for currentCell in currentRow:
                    self.combobox_profile.addItem(currentCell.value)



    def set_values_cb_grade_steel(self, type_profile):
        self.combobox_grade_steel.clear()
        if type_profile == "Двутавр":

            sheet_steel_grade = self.wb['Сталь Двутавр']
            max_row =sheet_steel_grade.max_row
            steel_list = []
            for currentRow in sheet_steel_grade[f'A3:A{max_row}']:
                for currentCell in currentRow:
                    if currentCell.value not in steel_list:
                        steel_list.append(currentCell.value)
            self.combobox_grade_steel.addItems(steel_list)


        elif type_profile == "Труба":

            sheet_steel_grade = self.wb['Сталь пластин']
            max_row = sheet_steel_grade.max_row
            steel_list = []
            for currentRow in sheet_steel_grade[f'A3:A{max_row}']:
                for currentCell in currentRow:
                    if currentCell.value not in steel_list:
                        steel_list.append(currentCell.value)
            self.combobox_grade_steel.addItems(steel_list)


    def get_output(self):
        type_profile = self.combobox_type_profile.currentText()
        profile = self.combobox_profile.currentText()
        grade_steel = self.combobox_grade_steel.currentText()
        text = f'Тип профиля: {type_profile}\nПрофиль: {profile}\nСталь: {grade_steel}'
        self.output.setText(text)


    def closeEvent(self, a0):
        self.wb.close()

app = QApplication(sys.argv)
window = Mainwindow()
window.show()

app.exec()